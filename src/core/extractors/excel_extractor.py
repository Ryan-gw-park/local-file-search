"""
Local Finder X v2.0 - Excel Extractor

Extracts content from .xlsx files using openpyxl.
Converts to Markdown table format for better searchability.
"""

from typing import List, Dict, Any

from .base import BaseExtractor, ExtractorResult, register_extractor

try:
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    load_workbook = None


@register_extractor
class ExcelExtractor(BaseExtractor):
    """Extractor for Microsoft Excel documents (.xlsx)."""
    
    SUPPORTED_EXTENSIONS = [".xlsx"]
    
    # Maximum rows per sheet to extract
    MAX_ROWS = 1000
    # Maximum columns to extract
    MAX_COLS = 50
    
    def extract(self, file_path: str) -> ExtractorResult:
        """Extract content from an Excel workbook."""
        if not XLSX_AVAILABLE:
            return self._create_error_result(
                "openpyxl is not installed. Install with: pip install openpyxl"
            )
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            sections = []
            full_text_parts = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_content = self._extract_sheet(sheet, sheet_name)
                
                if sheet_content["text"]:
                    sections.append({
                        "type": "sheet",
                        "name": sheet_name,
                        "content": [sheet_content["text"]],
                        "row_count": sheet_content["row_count"],
                    })
                    
                    # Add sheet header for context
                    full_text_parts.append(f"## Sheet: {sheet_name}")
                    full_text_parts.append(sheet_content["text"])
            
            wb.close()
            
            full_text = "\n\n".join(full_text_parts)
            
            metadata = {
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
            }
            
            return self._create_success_result(
                text=full_text,
                sections=sections,
                metadata=metadata,
            )
            
        except Exception as e:
            return self._create_error_result(f"Error extracting Excel file: {str(e)}")
    
    def _extract_sheet(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Extract content from a single sheet."""
        rows_data = []
        row_count = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(max_row=self.MAX_ROWS, max_col=self.MAX_COLS)):
            if row_idx >= self.MAX_ROWS:
                break
            
            # Get cell values
            cells = []
            has_content = False
            for cell in row:
                value = cell.value
                if value is not None:
                    has_content = True
                    cells.append(str(value).strip())
                else:
                    cells.append("")
            
            if has_content:
                rows_data.append(cells)
                row_count += 1
        
        if not rows_data:
            return {"text": "", "row_count": 0}
        
        # Convert to Markdown table format
        md_lines = []
        
        # Header row
        if rows_data:
            header = rows_data[0]
            md_lines.append("| " + " | ".join(header) + " |")
            md_lines.append("| " + " | ".join(["---"] * len(header)) + " |")
            
            # Data rows
            for row in rows_data[1:]:
                # Pad row to match header length
                while len(row) < len(header):
                    row.append("")
                md_lines.append("| " + " | ".join(row[:len(header)]) + " |")
        
        return {
            "text": "\n".join(md_lines),
            "row_count": row_count,
        }


__all__ = ["ExcelExtractor", "XLSX_AVAILABLE"]
